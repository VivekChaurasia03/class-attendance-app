#!/usr/bin/env python3
"""
Zoom Attendance Analyzer
========================
Usage:
    python analyze_attendance.py                     # interactive: pick from raw_reports/
    python analyze_attendance.py path/to/file.csv    # direct file
    python analyze_attendance.py --all               # process all CSVs in raw_reports/
"""

import os
import sys
import glob
import argparse
import pandas as pd
from datetime import datetime
from pathlib import Path

# ─────────────────────────────────────────────
# CONFIGURATION — edit these to match your class
# ─────────────────────────────────────────────
CONFIG = {
    "professor": "beth bonsignore",          # exact name (case-insensitive)
    "exclude": ["beth bonsignore", "jay walter", "Vivek Chaurasia"],  # professor + TAs to exclude
    "min_attendance_pct": 50,                # % of class time = "present"
    "late_threshold_min": 10,                # minutes after class start = "late"
    "early_leave_threshold_min": 10,         # minutes before class end = "left early"
    "ghost_join_max_min": 2,                 # sessions ≤ this are Zoom glitches
    "raw_reports_dir": "raw_reports",        # folder with input CSVs
    "output_dir": "attendance_reports",      # folder for output Excel files
}
# ─────────────────────────────────────────────


def parse_args():
    parser = argparse.ArgumentParser(description="Analyze Zoom attendance CSV reports")
    parser.add_argument("file", nargs="?", help="Path to a specific CSV file")
    parser.add_argument("--all", action="store_true", help="Process all CSVs in raw_reports/")
    return parser.parse_args()


def pick_file_interactively(raw_dir):
    files = sorted(glob.glob(os.path.join(raw_dir, "*.csv")))
    if not files:
        print(f"❌ No CSV files found in '{raw_dir}/'")
        print(f"   Put your Zoom attendance CSVs there and try again.")
        sys.exit(1)

    print("\n📂 Available attendance reports:")
    for i, f in enumerate(files, 1):
        size = os.path.getsize(f)
        mtime = datetime.fromtimestamp(os.path.getmtime(f)).strftime("%Y-%m-%d")
        print(f"  [{i}] {os.path.basename(f)}  ({size} bytes, modified {mtime})")

    print(f"  [A] Process ALL files")
    print()
    choice = input("Select a file number (or A for all): ").strip().lower()

    if choice == "a":
        return files
    try:
        idx = int(choice) - 1
        if 0 <= idx < len(files):
            return [files[idx]]
    except ValueError:
        pass

    print("❌ Invalid choice.")
    sys.exit(1)


def analyze_file(csv_path, config):
    print(f"\n{'='*60}")
    print(f"📋 Processing: {os.path.basename(csv_path)}")
    print(f"{'='*60}")

    # Load
    df = pd.read_csv(csv_path)
    df.columns = df.columns.str.strip()
    df['Join time'] = pd.to_datetime(df['Join time'])
    df['Leave time'] = pd.to_datetime(df['Leave time'])
    df['Duration(Minutes)'] = pd.to_numeric(df['Duration(Minutes)'], errors='coerce')

    # Get class bounds from professor row
    prof_rows = df[df['User Name'].str.lower() == config['professor'].lower()]
    if prof_rows.empty:
        print(f"⚠️  Professor '{config['professor']}' not found. Using first/last timestamps.")
        class_start = df['Join time'].min()
        class_end = df['Leave time'].max()
    else:
        class_start = prof_rows['Join time'].min()
        class_end = prof_rows['Leave time'].max()

    class_duration = (class_end - class_start).total_seconds() / 60
    date_str = class_start.strftime("%Y-%m-%d")
    time_str = class_start.strftime("%I:%M %p")

    print(f"📅 Date: {class_start.strftime('%B %d, %Y')}  |  "
          f"Time: {time_str}  |  Duration: {class_duration:.0f} min")

    # Filter out professor/TAs
    exclude_lower = [e.lower() for e in config['exclude']]
    students = df[~df['User Name'].str.lower().isin(exclude_lower)].copy()

    # Aggregate multiple sessions per student
    def agg_student(grp):
        email = next((v for v in grp['User Email'] if pd.notna(v) and str(v).strip()), '')
        total_min = grp['Duration(Minutes)'].sum()
        sessions = len(grp)
        first_join = grp['Join time'].min()
        last_leave = grp['Leave time'].max()
        # Exclude ghost sessions (≤ ghost_join_max_min) from "real" sessions count
        real_sessions = grp[grp['Duration(Minutes)'] > config['ghost_join_max_min']]
        real_session_count = len(real_sessions)
        return pd.Series({
            'Email': email,
            'Total_Minutes': round(total_min, 1),
            'Raw_Sessions': sessions,
            'Real_Sessions': real_session_count,
            'First_Join': first_join,
            'Last_Leave': last_leave,
        })

    summary = students.groupby('User Name', group_keys=False).apply(agg_student).reset_index()

    # Derived columns
    summary['Pct_Attended'] = (summary['Total_Minutes'] / class_duration * 100).round(1)
    summary['Late_Join_Min'] = ((summary['First_Join'] - class_start)
                                 .dt.total_seconds() / 60).round(1).clip(lower=0)
    summary['Left_Early_Min'] = ((class_end - summary['Last_Leave'])
                                  .dt.total_seconds() / 60).round(1).clip(lower=0)
    summary['Is_Late'] = summary['Late_Join_Min'] > config['late_threshold_min']
    summary['Left_Early'] = summary['Left_Early_Min'] > config['early_leave_threshold_min']
    summary['Reconnected'] = summary['Real_Sessions'] > 1
    summary['Ghost_Only'] = summary['Total_Minutes'] <= config['ghost_join_max_min']

    summary['Status'] = summary.apply(lambda r:
        'Ghost Join' if r['Ghost_Only'] else
        ('Present' if r['Pct_Attended'] >= config['min_attendance_pct'] else 'Absent/Brief'),
        axis=1
    )

    # Flags column (human-readable)
    def build_flags(r):
        flags = []
        if r['Is_Late']:
            flags.append(f"late +{r['Late_Join_Min']:.0f}min")
        if r['Left_Early']:
            flags.append(f"left early -{r['Left_Early_Min']:.0f}min")
        if r['Reconnected']:
            flags.append(f"reconnected {r['Real_Sessions']}x")
        if r['Ghost_Only']:
            flags.append("ghost join (≤2min)")
        return ", ".join(flags)

    summary['Flags'] = summary.apply(build_flags, axis=1)

    # Print summary
    present = summary[summary['Status'] == 'Present'].sort_values('First_Join')
    absent = summary[summary['Status'] == 'Absent/Brief'].sort_values('Total_Minutes', ascending=False)
    ghosts = summary[summary['Status'] == 'Ghost Join']

    print(f"\n✅ Present: {len(present)}  |  ❌ Absent/Brief: {len(absent)}  |  "
          f"👻 Ghost Joins: {len(ghosts)}  |  Total unique: {len(summary)}\n")

    if len(absent) > 0:
        print("❌ ABSENT / BRIEF (below 50% threshold):")
        for _, r in absent.iterrows():
            print(f"   {r['User Name']:<35} {r['Total_Minutes']:.0f}min "
                  f"({r['Pct_Attended']:.0f}%)  {r['Flags']}")

    if len(ghosts) > 0:
        print("\n👻 GHOST JOINS (likely Zoom glitch, ≤2 min total):")
        for _, r in ghosts.iterrows():
            print(f"   {r['User Name']:<35} {r['Total_Minutes']:.0f}min")

    # ── Save Excel ──
    os.makedirs(config['output_dir'], exist_ok=True)
    out_filename = f"attendance_{date_str}_{class_start.strftime('%H%M')}.xlsx"
    out_path = os.path.join(config['output_dir'], out_filename)

    col_order = ['User Name', 'Email', 'Status', 'Total_Minutes', 'Pct_Attended',
                 'Real_Sessions', 'First_Join', 'Last_Leave', 'Is_Late',
                 'Left_Early', 'Reconnected', 'Flags']

    with pd.ExcelWriter(out_path, engine='openpyxl') as writer:
        from openpyxl.styles import PatternFill, Font, Alignment
        from openpyxl.utils import get_column_letter

        def write_sheet(df_sheet, sheet_name):
            df_sheet[col_order].to_excel(writer, sheet_name=sheet_name, index=False)
            ws = writer.sheets[sheet_name]
            # Style header
            for cell in ws[1]:
                cell.font = Font(bold=True)
                cell.fill = PatternFill("solid", fgColor="1F4E79")
                cell.font = Font(bold=True, color="FFFFFF")
            # Color rows by status
            status_colors = {
                'Present': 'E2EFDA',
                'Absent/Brief': 'FCE4D6',
                'Ghost Join': 'F2F2F2',
            }
            status_col = col_order.index('Status') + 1
            for row in ws.iter_rows(min_row=2):
                status_val = row[status_col - 1].value
                color = status_colors.get(status_val, 'FFFFFF')
                for cell in row:
                    cell.fill = PatternFill("solid", fgColor=color)
            # Auto-width
            for col_cells in ws.columns:
                max_len = max((len(str(c.value or '')) for c in col_cells), default=10)
                ws.column_dimensions[get_column_letter(col_cells[0].column)].width = min(max_len + 2, 40)

        # All students
        all_sorted = summary.sort_values(['Status', 'Pct_Attended'], ascending=[True, False])
        write_sheet(all_sorted, 'All Students')

        # Present
        write_sheet(present, 'Present')

        # Absent / Brief
        if len(absent) > 0:
            write_sheet(absent, 'Absent or Brief')

        # Ghost joins
        if len(ghosts) > 0:
            write_sheet(ghosts, 'Ghost Joins')

        # Raw data
        df.to_excel(writer, sheet_name='Raw Data', index=False)

        # Stats sheet
        stats = pd.DataFrame([
            ['Date', class_start.strftime('%B %d, %Y')],
            ['Class Start', class_start.strftime('%I:%M %p')],
            ['Class End', class_end.strftime('%I:%M %p')],
            ['Duration (min)', f'{class_duration:.0f}'],
            ['Total Unique Students', len(summary)],
            ['Present', len(present)],
            ['Absent/Brief', len(absent)],
            ['Ghost Joins', len(ghosts)],
            ['Attendance Rate', f'{len(present)/max(len(summary)-len(ghosts),1)*100:.1f}%'],
        ], columns=['Metric', 'Value'])
        stats.to_excel(writer, sheet_name='Class Stats', index=False)
        ws = writer.sheets['Class Stats']
        for cell in ws[1]:
            cell.font = Font(bold=True)
            cell.fill = PatternFill("solid", fgColor="1F4E79")
            cell.font = Font(bold=True, color="FFFFFF")
        for col_cells in ws.columns:
            max_len = max((len(str(c.value or '')) for c in col_cells), default=10)
            ws.column_dimensions[get_column_letter(col_cells[0].column)].width = min(max_len + 2, 40)

    print(f"\n💾 Report saved → {out_path}")
    return out_path


def main():
    args = parse_args()
    config = CONFIG

    # Determine which files to process
    if args.file:
        files = [args.file]
    elif args.all:
        files = sorted(glob.glob(os.path.join(config['raw_reports_dir'], "*.csv")))
        if not files:
            print(f"❌ No CSV files found in '{config['raw_reports_dir']}/'")
            sys.exit(1)
        print(f"📂 Found {len(files)} file(s) to process.")
    else:
        files = pick_file_interactively(config['raw_reports_dir'])

    for f in files:
        analyze_file(f, config)

    print(f"\n✅ Done! Check the '{config['output_dir']}/' folder for your reports.\n")


if __name__ == "__main__":
    try:
        main()
    except KeyboardInterrupt:
        print("\n⚠️  Cancelled.")
        sys.exit(0)
